#!/usr/bin/env python
# coding: utf-8

# In[ ]:


"""
Project Name: B-Form Lite
Author: ASSR
License: GNU General Public License v3.0 (GPLv3)
See LICENSE file for details.

Copyright (C) 2025 ASSR
The program and its associated resources (images, Excel files) are licensed
under the GNU General Public License v3.0.

"""




import pandas as pd
import os
import xlwings as xw
import xlwings
import openpyxl
import shutil
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

def select_files_and_process(root, max_rows_per_sheet):
    files_selected = filedialog.askopenfilenames()
    if files_selected:
        selected_files = root.tk.splitlist(files_selected)
        output_folder = select_output_folder(root)
        if output_folder:
            for selected_file in selected_files:  # Iterate over each selected file
                process_file(selected_file, output_folder, max_rows_per_sheet)  # Call process_file() for each file
    else:
        messagebox.showerror("Error", "No files selected.")
    
def convert_to_xlsx(selected_file, output_folder):
    """
    Converts HTML, HTM, or XLS files to XLSX format.
    """
    file_extension = os.path.splitext(selected_file)[1].lower()
    new_filename = os.path.join(output_folder, f"converted_{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}.xlsx")
    
    if file_extension in [".xls"]:
        wb = xw.Book(selected_file)
        wb.save(new_filename)
        wb.close()
        print(f"Converted {selected_file} to {new_filename}")
    
    elif file_extension in [".html", ".htm"]:
        df = pd.read_html(selected_file)[0]  # Read the first table found
        df.to_excel(new_filename, index=False)
        print(f"Converted {selected_file} to {new_filename}")
    
    else:
        new_filename = selected_file  # If already XLSX, use the same file
    
    return new_filename

def process_file(selected_file, output_folder, max_rows_per_sheet):
    """
    Process the contents of the selected file and create a new Excel file with the same contents.
    """      
        
    # Define new_filename and create a new Excel file
    selected_file = convert_to_xlsx(selected_file, output_folder)
    
    timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    output_filename = os.path.join(output_folder, f"output_{timestamp}.xlsx")
        
    # Load the Excel file
    wb = openpyxl.load_workbook(selected_file)
    ws = wb.active
    
    # Create a new Excel workbook and worksheet
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
        
    # Copy the contents of the selected file to the new workbook
    for row in ws.iter_rows():
        new_ws.append([cell.value for cell in row])

        # Add a split command to split the headers
        split_command = "split"  # Modify this with your actual split command
        
    if new_ws["A2"].value:
        for index, character in enumerate(new_ws["A2"].value):
            new_ws.cell(row=2, column=index + 1).value = character
            
    if new_ws["D5"].value:
        for index, character in enumerate(new_ws["D5"].value):
            new_ws.cell(row=3, column=index + 1).value = character

    # Dictionary mapping codes to full forms
    code_to_full_form = {"IC": "ICEAS"}

    # Save the new Excel file
    # Ensure new_filename is properly assigned
    new_filename = output_filename 

    # Save the new Excel file
    new_wb.save(new_filename)

    print(f"Output file created: {new_filename}")
        
    # Now we have the new_filename, let's call process_excel_file with it
    process_excel_file(new_filename, max_rows_per_sheet)

def process_excel_file(new_filename, max_rows_per_sheet):
    # Implement your logic to process the Excel file here
    print(f"Processing Excel file: {new_filename}")
    
    def find_file(filename):
        """Search for the file anywhere in the system and return its path."""
        for root, _, files in os.walk("C:\\"):  # Change "C:\\" to "/" for Linux/Mac
            if filename in files:
                 return os.path.join(root, filename)
        return None

    # Search for the required files
    image_path = find_file("vtu.png")
    excel_template_path = find_file("empty_b_form.xlsx")

    if not image_path or not excel_template_path:
        print("Error: Required files (image and/or Excel template) not found on the system!")
        return

    print(f"Image found at: {image_path}")
    print(f"Excel template found at: {excel_template_path}")
    
    class ExcelDataCopier:
        def __init__(self, source_file, excel_template_path, max_rows_per_sheet, image_path):
            self.source_file = source_file
            self.excel_template_path = excel_template_path
            self.source_wb = openpyxl.load_workbook(source_file)
            self.dest_wb = openpyxl.load_workbook(excel_template_path)
            self.max_rows_per_sheet = max_rows_per_sheet
            self.image_path = image_path  # Store image path

        def copy_data(self, source_cells, dest_cells):
            source_ws = self.source_wb.active
            dest_ws = self.dest_wb.active

            for source_cell, dest_cell in zip(source_cells, dest_cells):
                dest_ws[dest_cell].value = source_ws[source_cell].value
                
        def copy_values_with_images(self):   
            source_ws = self.source_wb.active
            dest_ws = self.dest_wb.active
            # Variables for tracking the rows in the destination sheet
            current_row = 17 # Start copying to B17
            destination_sheet_index = 1

            # Extract the cell value to be used as the sheet name
            sheet_name = source_ws["H5"].value  # Change "H5" to the cell containing the desired sheet name
            title=f"{sheet_name}"
            dest_ws.title = title

            # Copy values from source to destination
            for row_index in range(5, source_ws.max_row + 1):  # Start from row 5 in the source sheet
                # Get the value from column B in the source sheet
                source_value = source_ws.cell(row=row_index, column=2).value
                # Copy the value to column B in the destination sheet
                dest_ws.cell(row=current_row, column=2).value = source_value

                # Move to the next row in the destination sheet
                current_row += 1

                # If 18 rows are copied, create a new sheet in the destination workbook
                if current_row > 16 + max_rows_per_sheet:
                    # Copy the first row of each set of 18 rows to cell C1
                    source_first_row_value = source_ws.cell(row=row_index - max_rows_per_sheet + 1, column=2).value
                    dest_ws.cell(row=12, column=7).value = source_first_row_value
                    # Copy the last row of each set of 18 rows to cell C2
                    source_last_row_value = source_ws.cell(row=row_index, column=2).value
                    dest_ws.cell(row=12, column=11).value = source_last_row_value
     
                    dest_ws = self.dest_wb.copy_worksheet(dest_ws)
                    img = Image(image_path)
                    img.height=54
                    img.width=676
                    dest_ws.add_image(img,"B2")
                    
                    title=f"{sheet_name}"
                    dest_ws.title = title
                    destination_sheet_index += 1
                    current_row = 17  # Reset current_row for the new sheet
                    for row in range(17, 43):
                        dest_ws[f"B{row}"].value = None
        
                        # Copy the first row of the last set of rows to cell C1
                        last_set_first_row_value = source_ws.cell(row=row_index + 1 , column=2).value
                        dest_ws.cell(row=12, column=7).value = last_set_first_row_value
                        # Copy the last row of the last set of rows to cell C2
                        last_set_last_row_value = source_ws.cell(row=source_ws.max_row, column=2).value
                        dest_ws.cell(row=12, column=11).value = last_set_last_row_value
            
        def save_excel_template_path(self, filename=None):
            source_ws = self.source_wb.active
            if filename is None:
                cell_name = source_ws["H5"].value  # Change "H5" to the cell containing the desired sheet name
                current_datetime = datetime.now()
                formatted_datetime = current_datetime.strftime("%d-%m-%Y_%H-%M-%S")
                filename = f"B-Form-{cell_name}-{formatted_datetime}.xlsx"
            self.dest_wb.save(filename)
            print(f"Destination file '{filename}' saved successfully.")
            
    # Call your actual processing function here
    copier = ExcelDataCopier(new_filename, excel_template_path, max_rows_per_sheet, image_path)
    
    source_cells = ["A3","B3","C3","D3","E3","F3","G3","H3","I3","F5","G5","I5","AJ2","AK2","AE2","AF2","AG2","AH2","E5"]  # Example source cells
    dest_cells = ["I10","J10","K10","L10","M10","N10","O10","P10","Q10","G6","C12","E8","L6","M6","N6","O6","P6","Q6","C9"]    # Example destination cells

    # Specify the source cells from which you want to copy the information
    source_date_cells = ["AM2","AN2","AL2","AJ2","AK2","AI2","AE2","AF2","AG2","AH2"] #DATE

    # Specify the destination cell where you want to paste the combined value
    dest_combined_cell = "C14"

    # Specify the source cells from which you want to copy the information
    source_time_cells = ["BH2","BI2","BJ2","BK2","BL2"] #TIME

    # Specify the destination cell where you want to paste the combined value
    dest_combined_cell_time = "I14"
    
    copier.copy_data(source_cells, dest_cells)
    copier.copy_data_merged_cells(source_date_cells, dest_combined_cell)
    copier.copy_data_merged_cells_time(source_time_cells, dest_combined_cell_time)
    copier.copy_values_with_images()
    copier.save_excel_template_path()

    # Specify the merged cells in the destination where you want to paste the value
    merged_dest_cells = ["C9:C10"]
    
def select_output_folder(root):
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        os.chdir(folder_selected)  # Set the selected folder as the working directory
        return folder_selected
    else:
        messagebox.showerror("Error", "No output folder selected.")
        return None

def main():
    root = tk.Tk()
    root.title("B-Form Lite")
    root.state('zoomed')  # Maximized window

    # Change background color for contrast
    root.configure(background='white')

    def set_background_image():
        try:
            background_image = tk.PhotoImage(file="")
            background_label = tk.Label(root, image=background_image)
            background_label.place(x=0, y=0, relwidth=1, relheight=1)  # Full window coverage
            background_label.image = background_image  # Keep a reference to prevent garbage collection
        except Exception as e:
            print("Error loading background image:", e)

    # Set background image
    set_background_image()

    def process_files():
        max_rows_per_sheet = max_rows_entry.get()
        if max_rows_per_sheet:
            select_files_and_process(root, int(max_rows_per_sheet))

    max_rows_label = tk.Label(root, text="Enter the maximum number of rows per sheet:")
    max_rows_label.pack(padx=100, pady=10)

    max_rows_entry = tk.Entry(root)
    max_rows_entry.pack(padx=100, pady=30)
    
    process_button = tk.Button(root, text="Select Files and Process", command=process_files, bg="white", fg="black", font=("Times New Roman", 12, "bold"), borderwidth=0, highlightthickness=0)
    process_button.pack(padx=100, pady=30)

    root.mainloop()
    
if __name__ == "__main__":
    main()


# In[ ]:




